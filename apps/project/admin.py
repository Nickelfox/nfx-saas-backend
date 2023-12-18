from django.contrib import admin
from django.shortcuts import render
from django.urls import path
from apps.project.models import Project, ProjectMember
from apps.project.resources import ProjectResource
from apps.team.models import Team
from apps.client.models import Client
from custom_admin import company_admin_site
from common.helpers import module_perm
from import_export.admin import ImportExportModelAdmin
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from common.constants import Color_choice, Project_type
from openpyxl.styles import PatternFill

# Register your models here.


class MemberInline(admin.TabularInline):
    model = ProjectMember
    fields = [
        "member",
    ]
    extra = 0
    can_delete = False
    show_change_link = True

    def get_queryset(self, request):
        user = request.user
        return (
            super()
            .get_queryset(request)
            .filter(member__company_id=user.company_id)
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "member":
            user = request.user
            kwargs["queryset"] = Team.objects.filter(
                company_id=user.company_id
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "view")

    def has_add_permission(self, request, obj=None):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "delete")


class ProjectSpecificAdmin(ImportExportModelAdmin):
    change_list_template = "company_admin/project_change_list.html"
    resource_class = ProjectResource
    inlines = [MemberInline]
    list_display = ["project_name", "client", "start_date", "end_date", "id"]
    list_filter = (
        "client",
        "project_type",
    )
    fields = [
        "project_name",
        "color_code",
        "project_code",
        "client",
        "start_date",
        "end_date",
        "project_type",
        "notes",
    ]

    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for instance in instances:
            instance.company_id = request.user.company_id
            instance.save()
        formset.save_m2m()

    def save_model(self, request, obj, form, change):
        # Save the object initially to generate obj.id
        super().save_model(request, obj, form, change)

        # Check if this is a new invitation being added (not an update)
        if not change:
            obj.company_id = request.user.company_id
            obj.save()

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "client":
            user = request.user
            kwargs["queryset"] = Client.objects.filter(
                company_id=user.company_id
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        user = request.user
        return super().get_queryset(request).filter(company_id=user.company_id)

    def get_urls(self):
        urls = super(ProjectSpecificAdmin, self).get_urls()
        my_urls = [
            path(
                "instruct-project/",
                self.admin_site.admin_view(self.instruct_view),
                name="project_template_instruct_view",
            ),
            path(
                "download_template/project/",
                self.admin_site.admin_view(self.download_template),
                name="project_download_template",
            ),
        ]
        return my_urls + urls

    def instruct_view(self, request):
        return render(
            request,
            "import_instruction/project_template.html",
        )

    def download_template(self, request):
        # If no items are selected, queryset will be None
        client_name = Client.objects.filter(
            company_id=request.user.company_id
        ).values_list("name", flat=True)
        # Create a workbook in-memory
        wb = Workbook()
        ws = wb.active
        headers = [
            "client",
            "project_name",
            "project_code",
            "start_date",
            "end_date",
            "color_code",
            "project_type",
            "notes",
        ]

        column_A_range = "A2:A1048576"
        column_D_range = "D2:D1048576"
        column_E_range = "E2:E1048576"
        column_F_range = "F2:F1048576"
        column_G_range = "G2:G1048576"
        ws.append(headers)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20

        choice_1_str = ",".join(client_name)
        choice_2_str = ",".join(Project_type.values)
        choice_3_str = ",".join(Color_choice.labels)
        valid_1_options = f'"{choice_1_str}"'
        valid_2_options = f'"{choice_2_str}"'
        valid_3_options = f'"{choice_3_str}"'

        required_column = ["A1", "B1", "F1"]
        fill_color = PatternFill(
            start_color="6fa8dc", end_color="6fa8dc", fill_type="gray125"
        )
        for cell in required_column:
            ws[cell].fill = fill_color

        rule = DataValidation(
            type="list",
            formula1=valid_1_options,
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True,
        )

        rule.error = "Entry not Valid"
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "please select from list"
        rule.promptTitle = "Select Option"
        ws.add_data_validation(rule)
        rule.add(column_A_range)

        rule = DataValidation(
            type="list",
            formula1=valid_2_options,
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True,
        )

        rule.error = "Entry not Valid"
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "please select from list"
        rule.promptTitle = "Select Option"
        ws.add_data_validation(rule)
        rule.add(column_G_range)

        rule = DataValidation(
            type="list",
            formula1=valid_3_options,
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True,
        )

        rule.error = "Entry not Valid"
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "please select from list"
        rule.promptTitle = "Select Option"
        ws.add_data_validation(rule)
        rule.add(column_F_range)

        rule = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )

        rule.error = (
            "Invalid date format or value suggested: (yyyy-mm-dd),(dd/mm/yyyy)"
        )
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "Enter a valid date supported: (yyyy-mm-dd),(dd/mm/yyyy)"
        rule.promptTitle = "Date Format"
        ws.add_data_validation(rule)
        rule.add(column_D_range)

        rule = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )

        rule.error = (
            "Invalid date format or value suggested: (yyyy-mm-dd),(dd/mm/yyyy)"
        )
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "Enter a valid date supported: (yyyy-mm-dd),(dd/mm/yyyy)"
        rule.promptTitle = "Date Format"
        ws.add_data_validation(rule)
        rule.add(column_E_range)

        # Create an HttpResponse with Excel content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set the response headers for file attachment
        response[
            "Content-Disposition"
        ] = "attachment; filename=project_template.xlsx"

        # Save the workbook directly to the response
        wb.save(response)

        return response

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project", user, "view")

    def has_add_permission(self, request):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project", user, "delete")


company_admin_site.register(Project, ProjectSpecificAdmin)


class ProjectMemberSpecificAdmin(admin.ModelAdmin):
    list_display = [
        "project",
        "member",
    ]
    list_filter = (
        "project",
        "member",
    )
    fields = [
        "project",
        "member",
    ]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        user = request.user
        if db_field.name == "project":
            model = Project
        if db_field.name == "member":
            model = Team
        kwargs["queryset"] = model.objects.filter(company_id=user.company_id)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        user = request.user
        return (
            super()
            .get_queryset(request)
            .filter(project__company_id=user.company_id)
        )

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "view")

    def has_add_permission(self, request):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("project_member", user, "delete")


company_admin_site.register(ProjectMember, ProjectMemberSpecificAdmin)
