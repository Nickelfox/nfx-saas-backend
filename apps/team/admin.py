from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from openpyxl import Workbook
from apps.department.models import Department
from apps.team.resources import TeamResource
from apps.user.models import User
from apps.project.models import Project, ProjectMember
from custom_admin import company_admin_site
from apps.team.models import Team
from common.helpers import module_perm
from django.contrib.admin.widgets import FilteredSelectMultiple
from common.constants import Days_choice
from django import forms
from datetime import timedelta
from import_export.admin import ImportExportModelAdmin
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill


# Register your models here.


class TeamSpecificAdminForm(forms.ModelForm):
    # Create a list of choices based on the Days_choice enumeration
    work_days = forms.MultipleChoiceField(
        choices=[(day.value, day.label) for day in Days_choice],
        widget=FilteredSelectMultiple("Work Days", is_stacked=False),
        required=False,
        initial=[
            "MON",
            "TUE",
            "WED",
            "THU",
            "FRI",
        ],  # Set default values for work_days
    )

    class Meta:
        model = Team
        fields = [
            "capacity",
            "emp_id",
            "work_days",
            "user",
            "department",
        ]

    def clean(self):
        cleaned_data = super().clean()
        # Set a default value for the "capacity" field if it's not provided
        if "capacity" not in cleaned_data:
            cleaned_data["capacity"] = timedelta(hours=8)
        return cleaned_data


class ProjectInline(admin.TabularInline):
    model = ProjectMember
    fields = [
        "project",
    ]
    extra = 0
    can_delete = False
    show_change_link = True

    def get_queryset(self, request):
        user = request.user
        return (
            super()
            .get_queryset(request)
            .filter(project__company_id=user.company_id)
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "project":
            user = request.user
            kwargs["queryset"] = Project.objects.filter(
                company_id=user.company_id
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "view")

    def has_add_permission(self, request, obj=None):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "delete")


class TeamSpecificAdmin(ImportExportModelAdmin):
    change_list_template = "company_admin/team_change_list.html"
    resource_class = TeamResource
    inlines = [ProjectInline]
    form = TeamSpecificAdminForm
    list_display = [
        "user",
        "department",
        "emp_id",
        "id",
        "capacity",
        "work_days",
    ]
    list_filter = ("department",)

    def save_model(self, request, obj, form, change):
        # Save the object initially to generate obj.id
        super().save_model(request, obj, form, change)

        # Check if this is a new invitation being added (not an update)
        if not change:
            obj.company_id = request.user.company_id
            obj.save()

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "department":
            user = request.user
            kwargs["queryset"] = Department.objects.filter(
                company_id=user.company_id
            )
        if db_field.name == "user":
            user = request.user
            kwargs["queryset"] = User.objects.filter(company=user.company_id)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        user = request.user
        return super().get_queryset(request).filter(company_id=user.company_id)

    def get_urls(self):
        urls = super(TeamSpecificAdmin, self).get_urls()
        my_urls = [
            path(
                "instruct-team/",
                self.admin_site.admin_view(self.instruct_view),
                name="team_template_instruct_view",
            ),
            path(
                "download_template/team/",
                self.admin_site.admin_view(self.download_template),
                name="team_download_template",
            ),
        ]
        return my_urls + urls

    def instruct_view(self, request):
        return render(
            request,
            "import_instruction/team_template.html",
        )

    def download_template(self, request):
        # If no items are selected, queryset will be None
        department_name = Department.objects.filter(
            company_id=request.user.company_id
        ).values_list("name", flat=True)
        # Create a workbook in-memory
        wb = Workbook()
        ws = wb.active
        headers = [
            "full_name",
            "email",
            "department",
            "designation",
            "emp_id",
        ]

        column_C_range = "C2:C1048576"
        ws.append(headers)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15

        choice_1_str = ",".join(department_name)
        valid_1_options = f'"{choice_1_str}"'

        required_column = ["A1", "B1", "C1"]
        fill_color = PatternFill(
            start_color="6fa8dc", end_color="6fa8dc", fill_type="gray125"
        )
        for cell in required_column:
            ws[cell].fill = fill_color

        rule = DataValidation(
            type="list",
            formula1=valid_1_options,
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True,
        )

        rule.error = "Entry not Valid"
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "please select from list"
        rule.promptTitle = "Select Option"
        ws.add_data_validation(rule)
        rule.add(column_C_range)

        # Create an HttpResponse with Excel content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set the response headers for file attachment
        response[
            "Content-Disposition"
        ] = "attachment; filename=team_template.xlsx"

        # Save the workbook directly to the response
        wb.save(response)

        return response

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "view")

    def has_add_permission(self, request):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("team", user, "delete")


company_admin_site.register(Team, TeamSpecificAdmin)
