from django.contrib import admin
from apps.project.models import ProjectMember
from apps.schedule.models import Schedule
from apps.team.models import Team
from apps.client.models import Client
from apps.project.models import Project, ProjectMember
from custom_admin import company_admin_site
from common.helpers import module_perm
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from openpyxl import Workbook
from import_export.admin import ImportExportModelAdmin
from openpyxl.worksheet.datavalidation import DataValidation
from apps.schedule.resources import ScheduleResource
from openpyxl.styles import PatternFill
# Register your models here.


class ScheduleSpecificAdmin(ImportExportModelAdmin):
    change_list_template = "company_admin/schedule_change_list.html"
    resource_class = ScheduleResource
    list_display = [
        "project_member",
        "start_at",
        "end_at",
        "id",
        "schedule_type",
    ]
    list_filter = (
        "start_at",
        "end_at",
    )
    fields = [
        "project_member",
        "start_at",
        "end_at",
        "notes",
        "assigned_hour",
        "schedule_type",
    ]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "project_member":
            user = request.user
            kwargs["queryset"] = ProjectMember.objects.filter(
                project__company_id=user.company_id
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_queryset(self, request):
        user = request.user
        return (
            super()
            .get_queryset(request)
            .filter(project_member__project__company_id=user.company_id)
        )
    
    def get_urls(self):
        urls = super(ScheduleSpecificAdmin, self).get_urls()
        my_urls = [
            path(
                "instruct-team/",
                self.admin_site.admin_view(self.instruct_view),
                name="schedule_template_instruct_view",
            ),
            path(
                "download_template/schedule/",
                self.admin_site.admin_view(self.download_template),
                name="schedule_download_template",
            ),
        ]
        return my_urls + urls
    
    def instruct_view(self, request):
        return render(
            request,
            "import_instruction/schedule_template.html",
        )
    
    def download_template(self, request):
        # Create a workbook in-memory
        wb = Workbook()
        ws = wb.active
        headers = [
            "project_name",
            "full_name",
            "start_at",
            "end_at",
            "assigned_hour",
            "notes",
        ]
        column_A_range = "A2:A1048576"
        column_B_range = "B2:B1048576"
        column_C_range = "C2:C1048576"
        column_D_range = "D2:D1048576"
        column_E_range = "E2:E1048576"
        column_F_range = "F2:F1048576"

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.append(headers)

        required_column = ["A1", "B1", "C1", "D1", "E1"]
        fill_color = PatternFill(
            start_color="6fa8dc", end_color="6fa8dc", fill_type="gray125"
        )
        for cell in required_column:
            ws[cell].fill = fill_color

        project_names = list(set(list(Project.objects.values_list('project_name', flat=True))))
        project_ids = list(ProjectMember.objects.values_list('project__id', flat=True))
        member_names = {}
        projects = Project.objects.filter(id__in=project_ids).distinct()
        for project in projects:
            members = ProjectMember.objects.filter(project=project)
            member_names[project.project_name] = list(members.values_list('member__user__full_name', flat=True))

        choice_project_str = ",".join(project_names)
        valid_project_options = f'"{choice_project_str}"'

        rule_project = DataValidation(
            type="list",
            formula1=valid_project_options,
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True,
        )

        rule_project.error = "Entry not Valid"
        rule_project.errorTitle = "Invalid Entry"

        rule_project.prompt = "Please select from the list"
        rule_project.promptTitle = "Select Project"

        ws.add_data_validation(rule_project)
        rule_project.add(column_A_range)

        rule = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )

        rule.error = (
            "Invalid date format or value suggested: (yyyy-mm-dd),(dd/mm/yyyy)"
        )
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "Enter a valid date supported: (yyyy-mm-dd),(dd/mm/yyyy)"
        rule.promptTitle = "Date Format"
        ws.add_data_validation(rule)
        rule.add(column_C_range)

        rule = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2000,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )

        rule.error = (
            "Invalid date format or value suggested: (yyyy-mm-dd),(dd/mm/yyyy)"
        )
        rule.errorTitle = "Invalid Entry"

        rule.prompt = "Enter a valid date supported: (yyyy-mm-dd),(dd/mm/yyyy)"
        rule.promptTitle = "Date Format"
        ws.add_data_validation(rule)
        rule.add(column_D_range)

        # Create an HttpResponse with Excel content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set the response headers for file attachment
        response[
            "Content-Disposition"
        ] = "attachment; filename=schedule_template.xlsx"

        # Save the workbook directly to the response
        wb.save(response)

        return response

    def has_change_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("schedule", user, "update")

    def has_view_permission(self, request, obj=None):
        # Check if the user has permission to change the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("schedule", user, "view")

    def has_add_permission(self, request):
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("schedule", user, "add")

    def has_delete_permission(self, request, obj=None):
        # Check if the user has permission to delete the object
        user = request.user
        if user.is_company_owner:
            return True
        else:
            return module_perm("schedule", user, "delete")


company_admin_site.register(Schedule, ScheduleSpecificAdmin)
